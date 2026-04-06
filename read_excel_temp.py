import sys

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook

wb = load_workbook("C:/Users/Matt/Lookitry/BASE_CLIENTES_CRM.xlsx")
ws = wb["BASE_CLIENTES"]

print("=== ANALISIS DE NOMBRE_EMPRESA (buscar nichos relevantes) ===")

# Palabras clave para negocio de moda/ropa/accesorios
moda_keywords = [
    "ROPA",
    "BOUTIQUE",
    "TIENDA",
    "TEXTIL",
    "VESTIR",
    "MODA",
    "ACESSORIOS",
    "ZAPATERIA",
    "CALZADO",
    "BILLETERA",
    "BOLSA",
    "CARTERA",
    "JOYA",
    "RELOJ",
    "FASHION",
    "STORE",
    "SHOP",
    "CLOTHING",
    "WEAR",
    "MODA",
]

# Todas las empresas
empresas = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[1]:  # NOMBRE_EMPRESA
        empresas.append(row[1].upper())

print(f"Total empresas: {len(empresas)}")

# Filtrar por keywords de moda
moda_count = 0
moda_empresas = []
for emp in empresas:
    for kw in moda_keywords:
        if kw in emp:
            moda_count += 1
            moda_empresas.append(emp)
            break

print(f"\nEmpresas relacionadas con moda/ropa: {moda_count}")
print("\n--- Muestra de empresas de moda ---")
for e in moda_empresas[:30]:
    print(f"  {e}")

print("\n\n=== EMPRESAS SIN keywords obvias (diversos) ===")
no_moda = [e for e in empresas if not any(kw in e for kw in moda_keywords)]
print(f"Total no-moda: {len(no_moda)}")
print("\n--- Muestra de empresas no-moda ---")
for e in no_moda[:50]:
    print(f"  {e}")
